#!/usr/bin/env python3
"""
OT Overview — Data Entry workbook  ->  website content.

Reads the "Data Entry" Excel workbook and regenerates ONE file:
    assets/js/content-overrides.js

That file is loaded last by index.html, so the workbook is the single source
of truth for all recap content (sessions + per-division Q&A). Curated visual /
structural data already in data.js (arcs, covenants, maps, etc.) is left alone.

USAGE
    python build_site_data.py                 # auto-find newest workbook beside this script
    python build_site_data.py "My Data Entry.xlsx"

Requires:  pip install openpyxl
"""
import sys, os, glob, json, datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency. Run:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))

# ---- where to read / write -------------------------------------------------
def find_workbook(arg):
    if arg:
        return arg if os.path.isabs(arg) else os.path.join(HERE, arg)
    cands = glob.glob(os.path.join(HERE, "*.xlsx"))
    cands = [c for c in cands if "~$" not in c]  # skip Excel lock files
    if not cands:
        sys.exit("No .xlsx workbook found next to this script. Pass the path as an argument.")
    cands.sort(key=os.path.getmtime, reverse=True)
    return cands[0]

OUT = os.path.join(HERE, "assets", "js", "content-overrides.js")

# ---- header -> field mapping (normalised header text) ----------------------
def norm(s):
    s = "" if s is None else str(s)
    out = []
    for ch in s.lower():
        out.append(ch if ch.isalnum() else " ")
    return " ".join("".join(out).split())

# exact (normalised) Sessions headers -> JS field
SESSION_MAP = {
    "session": "id",
    "book passage": "book",
    "chapter": "chapter",
    "topic": "topic",
    "starting question 3a": "recapQuestion",
    "recap answer 3b": "recapAnswer",
    "before you begin contexttone new": "contextTone",
    "passage divisions range title one per line": "divisions",
    "main point": "mainPoint",
    "key verse memory": "keyVerse",
    "thread god s kingdom": "kingdom",
    "thread god s salvation": "salvation",
    "thread god s promises": "promises",
    "god s character intent": "intention",
    "man s reality": "reality",
    "nt passage": "ntPassage",
    "nt point christ fulfilment": "ntPoint",
    "apply this week q1 ntapplication": "apply1",
    "apply this week q2": "apply2",
    "apply this week q3": "apply3",
    "cloze word ot": "clozeWord",
    "cloze context ot": "clozeContext",
    "cloze word nt": "clozeWordNt",
    "cloze context nt": "clozeContextNt",
    "further study links label url one per line": "furtherStudy",
    "highlight thread kingdom salvation promises primary focus": "highlight",
    "presenter leader note optional": "presenterNote",
    "pray for": "prayFor",
    "pause pray adore prompt": "actsAdore",
    "pause pray confess prompt": "actsConfess",
    "pause pray thank prompt": "actsThank",
    "pause pray supplicate prompt": "actsSupplicate",
    "flashcard hint main point hintmainpoint": "hintMainPoint",
    "flashcard hint threads hintthreads": "hintThreads",
    "flashcard hint tension hinttension": "hintTension",
    "flashcard hint nt christ fulfilment hintnt": "hintNt",
    "flashcard hint apply to me hintapply": "hintApply",
    "session type study review": "sessionType",
    "reviews sessions e g 2 11": "reviewsSessions",
}
# keyword fallbacks (so future columns work even if header wording shifts a bit)
def keyword_field(nh):
    if "meditate" in nh and "prompt" in nh: return "meditatePrompt"
    if "reflection" in nh and "focus" in nh: return "reflectionFocus"
    if "review" in nh and "question" in nh:  return "reviewQuestions"
    if "prayer" in nh and "prompt" in nh:    return "prayerPrompts"
    return None

# Division-question images that are not stored in the workbook — preserved here,
# keyed by (session#, division#, question#) using 1-based division/question nums.
PRESERVED_DIV_IMAGES = {
    (8, 2, 3):  ("imgSinaiVsTabernacle", "Recap-Dashboard/images/sinai-vs-tabernacle.png",
                 "Mt Sinai vs Tabernacle — visual comparison"),
    (10, 2, 2): ("imgTabernaclePhoto", "Recap-Dashboard/images/tabernacle-photo.png",
                 "Photo / diagram of the Tabernacle"),
}

def cell(v):
    if v is None: return ""
    s = str(v)
    return s.strip("\n").rstrip()

def read_sessions(ws):
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    col_field = {}
    for i, h in enumerate(header):
        nh = norm(h)
        f = SESSION_MAP.get(nh) or keyword_field(nh)
        if f: col_field[i] = f
    sessions = []
    for r in rows[1:]:
        if not r or cell(r[0]) == "": continue
        rec = {}
        for i, f in col_field.items():
            if i < len(r): rec[f] = cell(r[i])
        try: sid = int(float(rec.get("id", "")))
        except (ValueError, TypeError): continue
        rec["id"] = sid
        # applyQuestions array + ntApplication
        aq = [rec.pop("apply1", ""), rec.pop("apply2", ""), rec.pop("apply3", "")]
        aq = [a for a in aq if a]
        rec["applyQuestions"] = aq
        rec["ntApplication"] = aq[0] if aq else ""
        # highlight -> preserve the workbook value as-is (may be compound,
        # e.g. "Promises & Kingdom"); the dashboard handles these directly.
        rec["highlight"] = rec.get("highlight", "").strip()
        sessions.append(rec)
    sessions.sort(key=lambda x: x["id"])
    return sessions

def read_divisions(ws):
    rows = list(ws.iter_rows(values_only=True))
    by_session = {}
    for r in rows[1:]:
        if not r or cell(r[0]) == "": continue
        try: sid = int(float(cell(r[0])))
        except ValueError: continue
        def g(i): return cell(r[i]) if i < len(r) else ""
        try: dnum = int(float(g(2)))
        except ValueError: dnum = len(by_session.get(sid, [])) + 1
        rng, title = g(3), g(4)
        qs = []
        for qi, base in enumerate([5, 8, 11]):  # Q1/Q2/Q3 prompt,hint,answer
            prompt, hint, ans = g(base), g(base + 1), g(base + 2)
            if prompt:
                q = {"prompt": prompt, "hint": hint, "answer": ans}
                img = PRESERVED_DIV_IMAGES.get((sid, dnum, qi + 1))
                if img: q["__img"] = img
                qs.append(q)
        by_session.setdefault(sid, []).append({
            "dnum": dnum, "range": rng, "title": title,
            "questions": qs, "summary": g(14), "summaryHint": g(15),
        })
    for sid in by_session:
        by_session[sid].sort(key=lambda d: d["dnum"])
        for d in by_session[sid]: d.pop("dnum", None)
    return by_session

def js(v):
    return json.dumps(v, ensure_ascii=False)

def build_js(sessions, divisions, wb_name):
    # OT_WORKBOOK overlay (documented fields only; curated extras in data.js survive)
    wb_obj = {str(s["id"]): {k: v for k, v in s.items() if k != "id"} for s in sessions}
    # OT_DIVISION_DETAIL full object, resolving preserved images via window.__resources
    div_obj = {}
    img_lines = []
    for sid, divs in divisions.items():
        clean = []
        for d in divs:
            qs = []
            for q in d["questions"]:
                qq = {"prompt": q["prompt"], "hint": q["hint"], "answer": q["answer"]}
                if "__img" in q:
                    key, raw, alt = q["__img"]
                    qq["__IMG__"] = key      # placeholder, replaced below
                    qq["imageAlt"] = alt
                qs.append(qq)
            clean.append({"range": d["range"], "title": d["title"],
                          "questions": qs, "summary": d["summary"],
                          "summaryHint": d["summaryHint"]})
        div_obj[str(sid)] = clean

    body = []
    body.append("/* ════════════════════════════════════════════════════════════════")
    body.append("   AUTO-GENERATED — do not edit by hand.")
    body.append("   Source workbook : " + wb_name)
    body.append("   Generated       : " + datetime.datetime.now().strftime("%Y-%m-%d %H:%M"))
    body.append("   Regenerate with : python build_site_data.py")
    body.append("   This file is loaded LAST, so the workbook wins for every field below.")
    body.append("   ════════════════════════════════════════════════════════════════ */")
    body.append("(function(){")
    body.append("  var R = (typeof window !== 'undefined' && window.__resources) || {};")
    body.append("  var WB = " + js(wb_obj) + ";")
    body.append("  // Overlay workbook fields onto each session (workbook authoritative).")
    body.append("  if (window.OT_SESSIONS) {")
    body.append("    window.OT_SESSIONS.forEach(function(s){")
    body.append("      var o = WB[String(s.id)]; if (!o) return;")
    body.append("      for (var k in o) { s[k] = o[k]; }")
    body.append("    });")
    body.append("  }")
    # division detail with image resolution
    div_js = js(div_obj)
    body.append("  var DIV = " + div_js + ";")
    body.append("  // resolve preserved division images through window.__resources")
    body.append("  Object.keys(DIV).forEach(function(sid){")
    body.append("    DIV[sid].forEach(function(d){")
    body.append("      (d.questions||[]).forEach(function(q){")
    body.append("        if (q.__IMG__) { q.image = R[q.__IMG__] || ('assets/img/' + ({imgTabernaclePhoto:'tabernacle-photo',imgSinaiVsTabernacle:'sinai-vs-tabernacle'}[q.__IMG__]||'') + '.webp'); delete q.__IMG__; }")
    body.append("      });")
    body.append("    });")
    body.append("  });")
    body.append("  window.OT_DIVISION_DETAIL = DIV;")
    body.append("})();")
    return "\n".join(body) + "\n"

def main():
    wb_path = find_workbook(sys.argv[1] if len(sys.argv) > 1 else None)
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    if "Sessions" not in wb.sheetnames or "Division Questions" not in wb.sheetnames:
        sys.exit("Workbook must have 'Sessions' and 'Division Questions' sheets.")
    sessions = read_sessions(wb["Sessions"])
    divisions = read_divisions(wb["Division Questions"])
    out = build_js(sessions, divisions, os.path.basename(wb_path))
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        f.write(out)
    print("Read workbook : " + os.path.basename(wb_path))
    print("Sessions      : " + str(len(sessions)))
    print("Divisions     : " + str(sum(len(v) for v in divisions.values())) + " across " + str(len(divisions)) + " sessions")
    print("Wrote         : " + os.path.relpath(OUT, HERE))

if __name__ == "__main__":
    main()
